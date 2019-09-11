# Python script to use openpyxl module to work with excel spredsheets
# Reference: https://www.youtube.com/watch?v=q6Mc_sAPZ2Y

import openpyxl
import os

os.chdir(r'C:\Users\dhamodv\Documents\Data_Sets')

wb = openpyxl.load_workbook('australian_postcodes.xlsx')
#print(type(wb))
#print(wb.sheetnames)
sheet = wb['australian_postcodes']

# Extract the spreadsheet cell values as below
A1_Value = sheet['A1'].value
B1_Value = sheet['B2'].value
print(A1_Value, B1_Value)

# Update cell values as below
sheet['B2'].value = 'xxxxx'
print(sheet['B2'].value)
wb.save('australian_postcodes2.xlsx')

# Alternate way of reading cell values
A3_Value = sheet.cell(row=1, column=3).value
print(A3_Value)

# Reading more than one cell value
for i in range(1,5):
    print(sheet.cell(row=i, column=2).value)


# Update sheet title as below
Stitle = sheet.title
print(Stitle)
sheet.title = 'Updated Title'
wb.save('australian_postcodes2.xlsx')
